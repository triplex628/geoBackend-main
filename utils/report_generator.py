from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from django.http import HttpResponse
from api.models import TaskModel, EmployeeTaskModel, ItemModel, EmployeeModel
from openpyxl.utils import get_column_letter
import logging
from datetime import timedelta
from django.utils.timezone import now, localtime, is_naive, make_aware, get_current_timezone


def format_seconds(seconds):
    if seconds is None:
        return "0:00:00"
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"

class ReportGenerator:


    def generate_task_sheet(self, sheet, employee_tasks, end_time):
        """
        Генерация листа 'Отчет по задачам'
        """
        sheet.title = "Отчет по задачам"
        column_widths = [5, 30, 20, 20, 15, 20, 20, 30, 15]
        for col_num, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(col_num)].width = width

        headers = ["№", "Задача", "Дата постановки", "Дата завершения", "ID задачи", 
                    "Потраченное время", "Переделка", "Сотрудники", "Статус"]

        
        
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num).value = header
            sheet.cell(row=1, column=col_num).font = Font(bold=True)
            sheet.cell(row=1, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

        # Стартовая строка
        row = 2

        for task_index, task in enumerate(employee_tasks, start=1):
            
            task_title = task.task.title if task.task else "Не указано"
            start_time = task.start_time.strftime("%d.%m.%Y") if task.start_time else ""
            end_time = task.end_time.strftime("%d.%m.%Y") if task.end_time else "не завершена"
            total_time = format_seconds(task.total_time) if task.total_time else "0:00:00"
            rework_time = format_seconds(task.rework_time) if task.rework_time else "0:00:00"

            
            employees = []
            statuses = []

            for employee_task in EmployeeTaskModel.objects.filter(task=task.task):
                employee_name = f"{employee_task.employee.name} {employee_task.employee.surname}"
                employees.append(employee_name)

                # Определяем статус
                if employee_task.paused_message == "Ожидание по браку":
                    statuses.append("Закончено по браку")
                elif employee_task.paused_message == "Ожидание по комплектующим":
                    statuses.append("Закончена по недостатку комплектующих")
                elif employee_task.end_time:
                    statuses.append("Работа завершена")
                else:
                    statuses.append("В работе")

           
            sheet.cell(row=row, column=1).value = task_index  # №
            sheet.cell(row=row, column=2).value = task_title  # Задача
            sheet.cell(row=row, column=3).value = start_time  # Дата постановки
            sheet.cell(row=row, column=4).value = end_time  # Дата завершения
            sheet.cell(row=row, column=5).value = task.task.id if task.task else "N/A"  # ID задачи
            sheet.cell(row=row, column=6).value = total_time  # Потраченное время
            sheet.cell(row=row, column=7).value = rework_time  # Переделка
            sheet.cell(row=row, column=8).value = "\n".join(employees)  # Сотрудники
            sheet.cell(row=row, column=8).alignment = Alignment(wrap_text=True, vertical="top")
            sheet.cell(row=row, column=9).value = "\n".join(statuses)  # Статус
            sheet.cell(row=row, column=9).alignment = Alignment(wrap_text=True, vertical="top")

            row += 1





    def generate_equipment_sheet(self, sheet, start_time, end_time):
        """
        Генерация отчета 'По приборам' 
        """
        sheet.title = "По приборам"

        
        logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

        
        column_widths = [5, 15, 20, 20, 20, 30, 30, 25, 30]
        for col_num, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(col_num)].width = width

       
        headers = ["№", "ID прибора", "Дата начала работы", "Дата окончания работы", "Тип работы",
                "Продолжительность работы (общая)", "Переделка (кол-во времени)", "Сотрудник", "Комментарий"]

        try:
            
            employee_tasks = EmployeeTaskModel.objects.filter(
                start_time__gte=start_time,
                end_time__lte=end_time
            )
            logging.debug(f"Фильтрация задач успешна. Найдено {len(employee_tasks)} задач.")
        except Exception as e:
            logging.error(f"Ошибка при фильтрации задач: {e}")
            sheet.append(["Ошибка при фильтрации задач", str(e)])
            return

        
        tasks_by_item = {}
        for employee_task in employee_tasks:
            item_title = employee_task.item.title if employee_task.item else "Не указан"
            if item_title not in tasks_by_item:
                tasks_by_item[item_title] = []
            tasks_by_item[item_title].append(employee_task)

        # Стартовая строка
        row = 1

        for item_title, tasks in tasks_by_item.items():
            try:
                
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
                sheet.cell(row=row, column=1).value = f"Прибор: {item_title}"
                sheet.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=1).font = Font(bold=True, size=14)
                row += 1

                
                for col_num, header in enumerate(headers, start=1):
                    sheet.cell(row=row, column=col_num).value = header
                    sheet.cell(row=row, column=col_num).font = Font(bold=True)
                    sheet.cell(row=row, column=col_num).alignment = Alignment(horizontal="center", vertical="center")
                row += 1

                
                total_duration_seconds = sum(
                    [et.total_time if et.total_time else 0 for et in tasks],  
                    0
                )
                total_duration_hours = total_duration_seconds / 3600  # Преобразуем в часы
                logging.debug(f"Общая продолжительность для '{item_title}': {total_duration_hours:.2f} ч.")


                
                for task_index, employee_task in enumerate(tasks, start=1):
                    task = employee_task.task
                    item = employee_task.item
                    paused_message = employee_task.paused_message
                    is_finished = task.is_available

                    # Определяем комментарий
                    if paused_message == "Ожидание по браку":
                        comment = "Закончено по браку"
                    elif paused_message == "Ожидание по комплектующим":
                        comment = "Закончена по недостатку комплектующих"
                    elif end_time:
                        comment = "Работа завершена"
                    else:
                        comment = "В работе"

                    # Вычисляем безопасное время окончания задачи
                    end_times = [et.end_time for et in tasks if et.end_time]
                    end_time_task = max(end_times) if end_times else None

                    # Заполняем строку данными
                    sheet.cell(row=row, column=1).value = task_index  # Порядковый номер
                    sheet.cell(row=row, column=2).value = item.id if item else "Не указан"  # ID прибора
                    sheet.cell(row=row, column=3).value = employee_task.start_time.strftime("%d.%m.%Y") if employee_task.start_time else ""
                    sheet.cell(row=row, column=4).value = end_time_task.strftime("%d.%m.%Y") if end_time_task else "не завершена"
                    sheet.cell(row=row, column=5).value = task.get_type_of_task_display() if hasattr(task, 'get_type_of_task_display') else task.type_of_task  # Тип работы
                    sheet.cell(row=row, column=6).value = format_seconds(employee_task.total_time) if employee_task.total_time else "0:00:00" # Общая продолжительность
                    sheet.cell(row=row, column=7).value = format_seconds(employee_task.rework_time) if employee_task.rework_time else "0:00:00" # Время на переделку
                    sheet.cell(row=row, column=8).value = f"{employee_task.employee.name} {employee_task.employee.surname}"  # Сотрудник
                    sheet.cell(row=row, column=9).value = comment  # Комментарий
                    row += 1

                # отступ
                row += 5

                
                sheet.cell(row=row, column=1).value = "ИТОГ:"
                sheet.cell(row=row, column=1).font = Font(bold=True)
                sheet.cell(row=row, column=6).value = f"{total_duration_hours:.2f} ч"
                sheet.cell(row=row, column=6).font = Font(bold=True)
                row += 1

            except Exception as e:
                logging.error(f"Ошибка при обработке прибора '{item_title}': {e}")
                sheet.append([f"Ошибка при обработке прибора '{item_title}'", str(e)])
                continue







    def generate_employee_sheet(self, sheet, start_time, end_time):
        """
        Генерация третьего листа отчета с группировкой задач по сотрудникам.
        """
        sheet.title = "По сотрудникам"

        
        column_widths = [5, 15, 30, 20, 20, 20, 15, 20, 20]
        for col_num, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(col_num)].width = width

        
        headers = ["№", "ID Задачи", "Задача", "Дата взятия", "Дата окончания",
                "Потраченное время", "Переделка", "Статус работы"]
        
        
        row = 1

        
        employees = EmployeeModel.objects.all()

        for employee in employees:
            
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            sheet.cell(row=row, column=1).value = f"{employee.name} {employee.surname}"
            sheet.cell(row=row, column=1).font = Font(bold=True, size=14)
            sheet.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            row += 1

            
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=row, column=col_num).value = header
                sheet.cell(row=row, column=col_num).font = Font(bold=True)
                sheet.cell(row=row, column=col_num).alignment = Alignment(horizontal="center", vertical="center")
            row += 1

            
            employee_tasks = EmployeeTaskModel.objects.filter(
                employee=employee, start_time__gte=start_time, end_time__lte=end_time
            )

            if employee_tasks.exists():
                for task_index, task in enumerate(employee_tasks, start=1):
                    
                    sheet.cell(row=row, column=1).value = task_index  # №
                    sheet.cell(row=row, column=2).value = task.task.id if task.task else "N/A"  # ID Задачи
                    sheet.cell(row=row, column=3).value = task.task.title if task.task else "N/A"  # Задача
                    sheet.cell(row=row, column=4).value = task.start_time.strftime("%d.%m.%Y") if task.start_time else "N/A"  # Дата взятия
                    sheet.cell(row=row, column=5).value = task.end_time.strftime("%d.%m.%Y") if task.end_time else "N/A"  # Дата окончания
                    sheet.cell(row=row, column=6).value = format_seconds(task.total_time) if task.total_time else "0:00:00"  # Потраченное время
                    sheet.cell(row=row, column=7).value = format_seconds(task.rework_time) if task.rework_time else "0:00:00"  # Переделка
                    sheet.cell(row=row, column=8).value = self.get_task_status(task)  # Статус работы
                    row += 1
            else:
                
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
                sheet.cell(row=row, column=1).value = "Нет задач"
                sheet.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                row += 1

            
            row += 1

    def get_task_status(self, task):
        """
        Получение статуса работы.
        """
        if task.paused_message == "Ожидание по браку":
            return "Закончена по браку"
        elif task.paused_message == "Ожидание по комплектующим":
            return "Закончена по недостатку комплектующих"
        elif task.end_time:
            return "Работа завершена"
        else:
            return "В работе"






    def generate_report(self, start_time, end_time):
        print(f"Генерация отчета с {start_time} по {end_time}")
        workbook = Workbook()

        # Генерация первого листа
        sheet1 = workbook.active
        self.generate_task_sheet(sheet1, EmployeeTaskModel.objects.filter(start_time__gte=start_time, end_time__lte=end_time), end_time)


        # Генерация второго листа
        sheet2 = workbook.create_sheet("По приборам")
        try:
            self.generate_equipment_sheet(sheet2, start_time, end_time)
        except Exception as e:
            print(f"Ошибка в generate_equipment_sheet: {e}")

        # Генерация третьего листа
        sheet3 = workbook.create_sheet("По сотрудникам")
        self.generate_employee_sheet(sheet3, start_time, end_time)

        # Сохранение отчета
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=full_report.xlsx'
        workbook.save(response)
        return response

    


def generate_zalup_report(employee_tasks):
    """
    Генерация отчета на 1 листе.
    """

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Отчет по задачам"

    
    headers = [
        "№",
        "Сотрудник",
        "Задача",
        "ID задачи",
        "Дата постановки",
        "Дата запуска",
        "Дата завершения",
        "Полезное время",
        "Переделка",
        "Внерабочее время",
        "Общее время",
    ]
    sheet.append(headers)

    
    for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Текущая временная зона
    current_tz = get_current_timezone()
    current_time = now()  # Текущее время (осведомленное)
   
    for idx, employee_task in enumerate(employee_tasks, start=1):
        task = employee_task.task
        
        # Приведение start_time и end_time к осведомленному времени (если они наивные)
        start_time = employee_task.start_time
        if is_naive(start_time):
            start_time = make_aware(start_time, timezone=current_tz)

        end_time = employee_task.end_time
        if end_time and is_naive(end_time):
            end_time = make_aware(end_time, timezone=current_tz)

        created_at = task.created_at
        if is_naive(created_at):
            created_at = make_aware(created_at, timezone=current_tz)
        
        finished_at = task.finished_at
        if finished_at and is_naive(finished_at):
            finished_at = make_aware(finished_at, timezone=current_tz)
            

        if finished_at:
            total_seconds = employee_task.total_time  
        else:  
            total_seconds = int((current_time - localtime(task.created_at)).total_seconds())

        row = [
            idx,  # №
            f"{employee_task.employee.name} {employee_task.employee.surname}" if employee_task.employee else "Не указан",  # Сотрудник
            task.title if task else "Не указано",  # Задача
            task.id if task else "N/A",  # ID задачи
            task.created_at.strftime("%d.%m.%Y %H:%M:%S") if task and task.created_at else "Не указано",  # Дата постановки
            employee_task.start_time.strftime("%d.%m.%Y %H:%M:%S") if employee_task.start_time else "Не запущено",  # Дата запуска
            task.finished_at.strftime("%d.%m.%Y %H:%M:%S") if task.finished_at else "не завершена",  # Дата завершения
            format_seconds(employee_task.useful_time) if employee_task.useful_time else "0:00:00",  # Полезное время
            format_seconds(employee_task.rework_time) if employee_task.rework_time else "0:00:00",  # Переделка
            format_seconds(employee_task.non_working_time) if employee_task.non_working_time else "0:00:00",  # Внерабочее время
            format_seconds(employee_task.total_time) if employee_task.total_time else format_seconds(total_seconds),  # Общее время
        ]
        sheet.append(row)
    # Автоматическая подгонка ширины столбцов
    for col_num, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        for cell in column_cells:
            try:
                if cell.value:  #
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # немного пространства
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Выравниваем текст в ячейках
    # for col in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
    #     for cell in col:
    #         cell.alignment = Alignment(horizontal="left", vertical="top")

    # Генерация файла
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="task_report.xlsx"'
    wb.save(response)
    return response
